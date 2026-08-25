"""Format-specific exporters for the shared `Report` structure produced by
`services/reports.py` — one function per format, no per-format business
logic (every number is already computed by the time it reaches here).
"""
from __future__ import annotations

import csv
import io

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

BRAND = "ShopperMatch.AI"

# The funnel metrics (in order) that get a bar chart wherever the "Outreach
# Funnel" table appears — the rate rows (e.g. "Delivery Rate") are excluded
# since they're percentages, not counts, and would dwarf the chart's scale.
_FUNNEL_CHART_METRICS = {"Sent", "Delivered", "Opened", "Clicked", "Accepted", "Declined"}
_FUNNEL_CHART_COLORS = ["#0ea5e9", "#6366f1", "#8b5cf6", "#f59e0b", "#10b981", "#f43f5e"]


def _funnel_chart_rows(tables: list) -> list[tuple[str, int]] | None:
    for heading, _headers, rows in tables:
        if heading != "Outreach Funnel":
            continue
        return [(label, count) for label, count in rows if label in _FUNNEL_CHART_METRICS]
    return None


def _safe_sheet_title(name: str) -> str:
    # openpyxl sheet titles: <=31 chars, no []:*?/\\
    cleaned = "".join(c for c in name if c not in '[]:*?/\\')
    return cleaned[:31] or "Sheet"


def export_csv(report: dict) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([report["title"]])
    w.writerow([report["subtitle"]])
    w.writerow([f"Generated: {report['generated_at']}"])
    w.writerow([])

    w.writerow(["KPI", "Value"])
    for label, value in report["kpis"]:
        w.writerow([label, value])
    w.writerow([])

    for heading, rows in report.get("sections", []):
        w.writerow([heading])
        for label, value in rows:
            w.writerow([label, value])
        w.writerow([])

    for heading, headers, rows in report.get("tables", []):
        w.writerow([heading])
        w.writerow(headers)
        for row in rows:
            w.writerow(row)
        w.writerow([])

    return buf.getvalue().encode("utf-8")


def export_xlsx(report: dict) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    bold = Font(bold=True)

    summary.append([report["title"]])
    summary["A1"].font = Font(bold=True, size=14)
    summary.append([report["subtitle"]])
    summary.append([f"Generated: {report['generated_at']}"])
    summary.append([])

    summary.append(["KPI", "Value"])
    for cell in summary[summary.max_row]:
        cell.font = bold
    for label, value in report["kpis"]:
        summary.append([label, value])
    summary.append([])

    for heading, rows in report.get("sections", []):
        summary.append([heading])
        summary[summary.max_row][0].font = bold
        for label, value in rows:
            summary.append([label, value])
        summary.append([])

    for col in summary.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        summary.column_dimensions[col[0].column_letter].width = min(60, max(12, width + 2))

    for heading, headers, rows in report.get("tables", []):
        sheet = wb.create_sheet(_safe_sheet_title(heading))
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = bold
        for row in rows:
            sheet.append(row)
        for col in sheet.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(60, max(12, width + 2))

        if heading == "Outreach Funnel":
            # Relies on reports.py emitting the count rows (Sent..Declined) before
            # the percentage rows (Delivery/Open/Click Rate) — a contiguous run
            # from row 2, so a single Reference range can cover just the counts.
            chart_row_count = sum(1 for label, _count in rows if label in _FUNNEL_CHART_METRICS)
            if chart_row_count:
                chart = BarChart()
                chart.type = "col"
                chart.title = "Outreach Funnel"
                chart.y_axis.title = "Count"
                data_ref = Reference(sheet, min_col=2, min_row=1, max_row=1 + chart_row_count)
                cats_ref = Reference(sheet, min_col=1, min_row=2, max_row=1 + chart_row_count)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                chart.legend = None
                chart.height, chart.width = 9, 16
                sheet.add_chart(chart, "D2")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_pdf(report: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=18 * mm, bottomMargin=18 * mm, leftMargin=16 * mm, rightMargin=16 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(BRAND, styles["Heading4"]),
        Paragraph(report["title"], styles["Title"]),
        Paragraph(report["subtitle"] or "", styles["Normal"]),
        Paragraph(f"Generated {report['generated_at']}", styles["Normal"]),
        Spacer(1, 10 * mm),
    ]

    kpi_rows = [["KPI", "Value"]] + [[label, str(value)] for label, value in report["kpis"]]
    story.append(_styled_table(kpi_rows))
    story.append(Spacer(1, 6 * mm))

    chart_rows = _funnel_chart_rows(report.get("tables", []))
    if chart_rows:
        story.append(Paragraph("Outreach Funnel (chart)", styles["Heading3"]))
        story.append(_funnel_bar_chart(chart_rows))
        story.append(Spacer(1, 6 * mm))

    for heading, rows in report.get("sections", []):
        story.append(Paragraph(heading, styles["Heading3"]))
        story.append(_styled_table([["Field", "Value"]] + [[label, str(value) if value is not None else "—"] for label, value in rows]))
        story.append(Spacer(1, 6 * mm))

    for heading, headers, rows in report.get("tables", []):
        story.append(Paragraph(heading, styles["Heading3"]))
        if not rows:
            story.append(Paragraph("No data.", styles["Normal"]))
        else:
            story.append(_styled_table([headers] + [[str(c) for c in row] for row in rows]))
        story.append(Spacer(1, 6 * mm))

    doc.build(story)
    return buf.getvalue()


def _funnel_bar_chart(rows: list[tuple[str, int]]) -> Drawing:
    width, height = 160 * mm, 60 * mm
    drawing = Drawing(width, height)
    chart = VerticalBarChart()
    chart.x = 15 * mm
    chart.y = 12 * mm
    chart.width = width - 25 * mm
    chart.height = height - 20 * mm
    chart.data = [[count for _label, count in rows]]
    chart.categoryAxis.categoryNames = [label for label, _count in rows]
    chart.categoryAxis.labels.fontSize = 8
    chart.valueAxis.valueMin = 0
    chart.valueAxis.labels.fontSize = 8
    chart.barLabels.fontSize = 8
    chart.barLabelFormat = "%d"
    chart.barLabels.nudge = 8
    chart.groupSpacing = 10
    for i in range(len(rows)):
        chart.bars[(0, i)].fillColor = colors.HexColor(_FUNNEL_CHART_COLORS[i % len(_FUNNEL_CHART_COLORS)])
    drawing.add(chart)
    return drawing


def _styled_table(data: list[list]) -> Table:
    t = Table(data, hAlign="LEFT")
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return t


EXPORTERS = {"csv": export_csv, "xlsx": export_xlsx, "pdf": export_pdf}
MIME_TYPES = {
    "csv": "text/csv",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}
